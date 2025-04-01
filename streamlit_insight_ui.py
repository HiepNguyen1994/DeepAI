### GIAO DIỆN STREAMLIT

import streamlit as st
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import openpyxl
import requests
from bs4 import BeautifulSoup
from scipy.stats import zscore
from urllib.parse import urlparse
from googlesearch import search  # pip install googlesearch-python
from duckduckgo_search import DDGS  

st.set_page_config(layout="centered")

### TEST UI


st.title("💬 Banking AI")

# ====== Prompt hỏi bot ======
st.markdown("Nhập câu hỏi để truy vấn vào tài liệu ngân hàng")
query = st.text_input("", placeholder="Ví dụ: Tỷ lệ nợ xấu VPB năm 2024 là bao nhiêu?")

col1, col2 = st.columns([1, 1])
with col1:
    time_filter = st.selectbox("🗓️ Bộ lọc thời gian", ["mới nhất", "2024", "2023", "2022"], index=0)
with col2:
    source_filter = st.multiselect("🏩 Nguồn tài liệu", ["tất cả", "VDSC", "Vietcap", "HSC"], default=["tất cả"])
submit = st.button("🔍 Tìm kiếm")

# ====== TÍCH HỢP FAISS INSIGHT ======
from query_engine import get_answer

if submit and query:
    with st.spinner("🔎 Đang truy vấn từ kho tài liệu..."):
        try:
            insight_result = get_answer(
                doc_name=query,
                time_filter=time_filter,
                sources=source_filter,
                user_prompt=query
            )
            st.markdown("---")
            st.subheader("📄 Phân tích chuyên sâu như sau ")
            st.markdown(insight_result)
        except Exception as e:
            st.error(f"❌ Lỗi khi truy vấn FAISS Insight: {e}")

#### FAISS NEWS
from datetime import datetime, timedelta
from query_news import get_faiss_news_answer

st.title("📰 News chuyên sâu về ngân hàng")

col1, col2, col3 = st.columns([2, 1, 1])
with col1:
    query = st.text_input("Nhập truy vấn", placeholder="VD: VPBank tăng dự phòng năm 2024")

with col2:
    start_date = st.date_input("Từ ngày", value=datetime.today() - timedelta(days=30))

with col3:
    end_date = st.date_input("Đến ngày", value=datetime.today())

submit = st.button("🔍 Tìm trong kho dữ liệu")

if submit and query:
    with st.spinner("📰 Đang truy FAISS News..."):
        try:
            results = get_faiss_news_answer(
                query,
                start_date=start_date.strftime("%Y-%m-%d"),
                end_date=end_date.strftime("%Y-%m-%d")
            )
            st.markdown("---")
            st.subheader("📰 Tổng hợp tin tức gần đây")
            st.markdown(results)
        except Exception as e:
            st.error(f"❌ Lỗi FAISS News: {e}")


### SEARCH WEB
st.title("🔍 Search web để tìm tin mới nhất")

query = st.text_input("Nhập từ khóa tìm kiếm:", placeholder="Ví dụ: tỷ lệ nợ xấu VPBank 2024")
submit = st.button("🌐 Tìm kiếm trên web")

def get_page_summary_with_meta(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers, timeout=5)
        soup = BeautifulSoup(response.text, "html.parser")

        # Title
        title = soup.title.string.strip() if soup.title else url

        # Summary
        description = soup.find("meta", attrs={"name": "description"})
        summary = description["content"] if description and "content" in description.attrs else ""
        if not summary:
            first_p = soup.find("p")
            summary = first_p.get_text(strip=True) if first_p else ""

        # Source
        source = urlparse(url).netloc

        # Publish date
        if soup.find("time"):
            pub_time = soup.find("time").get("datetime", soup.find("time").text.strip())
        elif soup.find("meta", {"name": "pubdate"}):
            pub_time = soup.find("meta", {"name": "pubdate"}).get("content")
        elif soup.find("meta", {"property": "article:published_time"}):
            pub_time = soup.find("meta", {"property": "article:published_time"}).get("content")
        else:
            pub_time = "Không rõ thời gian"

        return title, summary, source, pub_time

    except Exception as e:
        return url, "Không thể tải nội dung.", "Không rõ nguồn", "Không rõ thời gian"

if submit and query:
    st.write(f"🧠 Kết quả tìm kiếm cho: **{query}**")
    with DDGS() as ddgs:
        results = ddgs.text(query, region="wt-wt", safesearch="moderate", max_results=5)
        for r in results:
            url = r["href"]
            title, summary, source, pub_time = get_page_summary_with_meta(url)
            st.markdown(f"### [{title}]({url})")
            st.markdown(f"📌 **Nguồn**: `{source}` | 🕒 **Thời gian**: {pub_time}")
            st.write(summary)
            st.markdown("---")

### VẼ CHART TỰ ĐỘNG

st.markdown("---")
st.subheader("📄 Tải file Excel/CSV và vẽ biểu đồ tự động")

uploaded_file = st.file_uploader("📂 Upload file Excel/CSV", type=['xlsx', 'xls', 'csv'])

def read_excel_with_dynamic_headers(uploaded_file, selected_sheet):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[selected_sheet]

    # Xử lý merged cells
    for merged_cell in list(ws.merged_cells):
        top_left_cell_value = ws.cell(merged_cell.min_row, merged_cell.min_col).value
        ws.unmerge_cells(str(merged_cell))
        for row in ws.iter_rows(min_row=merged_cell.min_row, max_row=merged_cell.max_row,
                                min_col=merged_cell.min_col, max_col=merged_cell.max_col):
            for cell in row:
                cell.value = top_left_cell_value

    df_raw = pd.DataFrame(ws.values)

    # Detect thời gian nằm ngang → transpose
    first_row = df_raw.iloc[0].astype(str)
    time_like = first_row.str.match(r"\d{4}-\d{2}-\d{2}|\d{4}/\d{2}/\d{2}|Q\d-\d{4}|^\d{4}$").sum()

    if time_like >= 3:
        df_raw = df_raw.T.reset_index(drop=True)
        st.info("🤖 Phát hiện thời gian nằm ngang → đã tự động xoay bảng.")

    st.markdown("🔍 Xem trước các dòng để chọn dòng làm tiêu đề:")
    preview = df_raw.head(10).fillna("").astype(str)
    st.dataframe(preview)

    header_row = st.selectbox("🧩 Chọn dòng làm header (theo chỉ số)", options=list(range(len(preview))), index=0)
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = df_raw.iloc[header_row].astype(str).str.strip().str.lower()
    df = df.loc[:, ~df.columns.duplicated()]
    df = df.replace(['None', 'nan'], np.nan).dropna(how='all').reset_index(drop=True)
    return df

# Hàm xử lý Outliers
def remove_outliers_iqr(df, cols):
    for col in cols:
        Q1 = df[col].quantile(0.25)
        Q3 = df[col].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - 1.5 * IQR
        upper_bound = Q3 + 1.5 * IQR
        df = df[(df[col] >= lower_bound) & (df[col] <= upper_bound)]
    return df

def remove_outliers_zscore(df, cols, threshold=3):
    return df[(np.abs(zscore(df[cols])) < threshold).all(axis=1)]


if uploaded_file:
    try:
        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            selected_sheet = st.selectbox("📁 Chọn sheet", wb.sheetnames)
            df = read_excel_with_dynamic_headers(uploaded_file, selected_sheet)

        st.success("✅ Đọc file thành công!")
        st.write("📄 Dữ liệu xem trước:", df.head())

        all_columns = df.columns.tolist()
        time_col = st.selectbox("🍓 Chọn cột thời gian/Ticker", all_columns)
        value_cols = st.multiselect("📊 Chọn cột số liệu", [c for c in all_columns if c != time_col])

        first_value_col = value_cols[0] if value_cols else ""
        unit_detected = "Giữ nguyên"
        if "usd" in first_value_col.lower() and "mn" in first_value_col.lower():
            unit_detected = "Triệu USD"
        elif "usd" in first_value_col.lower() and "bn" in first_value_col.lower():
            unit_detected = "Tỷ USD"
        elif any(x in first_value_col.lower() for x in ["triệu", "million", "mn"]):
            unit_detected = "Triệu VND"
        elif any(x in first_value_col.lower() for x in ["tỷ", "billion", "bn"]):
            unit_detected = "Tỷ VND"
        elif "usd" in first_value_col.lower():
            unit_detected = "USD"
        elif "%" in first_value_col or "tỷ lệ" in first_value_col.lower():
            unit_detected = "%"

        st.info(f"🤖 Gợi ý đơn vị hiện tại: {unit_detected}")

        units_options = ["Giữ nguyên", "VND", "Triệu VND", "Tỷ VND", "USD", "Triệu USD", "Tỷ USD", "%"]
        current_unit = st.selectbox("📌 Xác nhận đơn vị gốc", units_options, index=units_options.index(unit_detected) if unit_detected in units_options else 0)
        target_unit = st.selectbox("📀 Đơn vị hiển thị", units_options, index=units_options.index(unit_detected) if unit_detected in units_options else 0)

        exchange_rate = 24000
        if "USD" in [current_unit, target_unit]:
            exchange_rate = st.number_input("💱 Tỷ giá quy đổi USD → VND", value=24000)

        unit_factors = {
            "Giữ nguyên": 1,
            "VND": 1,
            "Triệu VND": 1e6,
            "Tỷ VND": 1e9,
            "USD": exchange_rate,
            "Triệu USD": 1e6 * exchange_rate,
            "Tỷ USD": 1e9 * exchange_rate,
            "%": 0.01
        }

        factor = unit_factors[current_unit] / unit_factors[target_unit]

    except Exception as e:
        st.error(f"❌ Lỗi khi xử lý file hoặc biểu đồ: {e}")

    # ========== Giao diện lọc và vẽ ==========
    data_type = st.radio("📂 Loại dữ liệu", ["Chuỗi thời gian (Time-Series)", "Danh mục độc lập (Categories)"], index=0)
    limit = st.slider("🔢 Số lượng dữ liệu muốn hiển thị", 5, 500, 30)

    sort_order = "Không sắp xếp"
    if data_type == "Chuỗi thời gian (Time-Series)":
        sort_order = st.radio("🔀 Sắp xếp theo thời gian", ["Tăng dần", "Giảm dần"], index=0)
    else:
        sort_order = st.radio("🔀 Sắp xếp dữ liệu theo giá trị", ["Không sắp xếp", "Tăng dần", "Giảm dần"], index=0)


    remove_outliers = st.checkbox("🚩 Loại bỏ giá trị ngoại lai (Outliers)", value=False)
    method = st.selectbox("🎯 Phương pháp Outlier", ["IQR", "Z-Score"], index=0) if remove_outliers else "IQR"

    def remove_outliers_iqr(df, cols):
        for col in cols:
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            df = df[(df[col] >= lower_bound) & (df[col] <= upper_bound)]
        return df

    def remove_outliers_zscore(df, cols, threshold=3):
        return df[(np.abs(zscore(df[cols])) < threshold).all(axis=1)]

    # Chuẩn bị dữ liệu
    plot_df = df[[time_col] + value_cols].dropna().copy()
    plot_df[value_cols] = plot_df[value_cols].apply(pd.to_numeric, errors='coerce') * factor

    if remove_outliers:
        original_len = len(plot_df)
        plot_df = remove_outliers_iqr(plot_df, value_cols) if method == "IQR" else remove_outliers_zscore(plot_df, value_cols)
        st.info(f"🧹 Đã loại bỏ {original_len - len(plot_df)} outliers bằng {method}.")

    # Slider chọn khoảng giá trị
    numeric_data = plot_df[value_cols].stack()
    min_val, max_val = float(numeric_data.min()), float(numeric_data.max())
    value_range = st.slider("📏 Chọn khoảng giá trị hiển thị", min_value=min_val, max_value=max_val, value=(min_val, max_val))

    plot_df = plot_df[(plot_df[value_cols[0]] >= value_range[0]) & (plot_df[value_cols[0]] <= value_range[1])]

    # Nút vẽ biểu đồ
    if st.button("📈 Vẽ biểu đồ"):
        if data_type.startswith("Chuỗi"):
            plot_df_final = (
                plot_df.sort_values(by=time_col, ascending=(sort_order == "Tăng dần"))
                .head(limit)  # ✅ Dùng head chứ không phải tail
                .reset_index(drop=True)
            )
        else:
            plot_df_final = plot_df
            if sort_order != "Không sắp xếp":
                plot_df_final = plot_df_final.sort_values(
                    by=value_cols[0], ascending=(sort_order == "Tăng dần")
                )
            plot_df_final = plot_df_final.head(limit).reset_index(drop=True)



        fig, ax1 = plt.subplots(figsize=(12, 6))
        ax2 = ax1.twinx() if len(value_cols) >= 2 else None

        colors = ['tab:blue', 'tab:orange', 'tab:green', 'tab:red']
        highlight_color = 'crimson'
        highlight_n = 3

        for i, col in enumerate(value_cols):
            ax = ax1 if i == 0 or not ax2 else ax2
            color = colors[i % len(colors)]

            x = plot_df_final[time_col].astype(str)
            y = plot_df_final[col]

            ax.plot(x, y, marker='o', label=col, color=color, linewidth=2.5, markersize=7)
            top_idx = y.nlargest(highlight_n).index
            ax.scatter(x.loc[top_idx], y.loc[top_idx], color=highlight_color, s=100, edgecolor='black', zorder=5)

        ax1.set_title("📊 Biểu đồ dữ liệu", fontsize=14)
        ax1.set_xlabel(time_col, fontsize=12)
        ax1.set_ylabel(value_cols[0], fontsize=12)
        if ax2 and len(value_cols) > 1:
            ax2.set_ylabel(value_cols[1], fontsize=12)

        for ax in [ax1, ax2] if ax2 else [ax1]:
            ax.grid(True, linestyle="--", alpha=0.5)
            ax.set_xticks(range(len(plot_df_final)))
            ax.set_xticklabels(plot_df_final[time_col].astype(str), rotation=45, ha='right')

        handles, labels = [], []
        for ax in [ax1, ax2] if ax2 else [ax1]:
            h, l = ax.get_legend_handles_labels()
            handles += h
            labels += l
        ax1.legend(handles, labels)

        st.pyplot(fig)





#### 



